#!/usr/bin/env python3
"""
SPU Log Analyzer — Complete Application
========================================
Single-file application that:
  1. Accepts raw SPU log files (CSV / XLSX, any common delimiter)
  2. Cleans & normalises the data automatically
  3. Lets you filter by time range via a modern GUI
  4. Generates a filtered data export (.xlsx)
  5. Generates a full fault analysis report (.xlsx) with:
       • Alarm tree with sub-event drill-down (MACs/IPs collapsed)
       • Class A/B/C/D highlights and actions
       • Hyperlinks between summary and detail sheets

Requirements:  pip install pandas openpyxl chardet
Run:           python spu_analyzer_complete.py
"""

import os, re, sys, threading, subprocess, io, traceback
from pathlib import Path
from datetime import datetime
from copy import deepcopy

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# ─── Optional chardet for encoding detection ──────────────────────────────────
try:
    import chardet
    HAS_CHARDET = True
except ImportError:
    HAS_CHARDET = False

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 1 — CLEANING ENGINE
# ══════════════════════════════════════════════════════════════════════════════

# Expected columns after cleaning (order matters for display)
CLEAN_COLS = ["Date", "Time", "Type", "Information", "Group", "Code",
              "Status", "Class", "Cause", "Consequence", "Action", "Parameters"]

# Aliases: maps raw column names (lower-stripped) → canonical name
COL_ALIASES = {
    "date": "Date", "datum": "Date",
    "time": "Time", "zeit": "Time",
    "timestamp": "Date", "datetime": "Date", "date_time": "Date",  # combined → Date, split later
    "type": "Type", "typ": "Type", "eventtype": "Type", "event_type": "Type",
    "information": "Information", "info": "Information", "description": "Information",
    "message": "Information", "msg": "Information", "text": "Information",
    "group": "Group", "grp": "Group", "category": "Group",
    "code": "Code", "alarm_code": "Code", "alarmcode": "Code",
    "status": "Status", "state": "Status",
    "class": "Class", "severity": "Class", "priority": "Class",
    "cause": "Cause", "reason": "Cause",
    "consequence": "Consequence", "effect": "Consequence",
    "action": "Action", "remedy": "Action", "recommendation": "Action",
    "parameters": "Parameters", "params": "Parameters", "details": "Parameters",
}

# Type value normalisation
TYPE_ALIASES = {
    "alarm": "alarm", "alm": "alarm", "alert": "alarm",
    "sysstatchangedonlinetime": "sysstatchangedonlinetime",
    "ycuinventory": "ycuinventory",
    "fspuchangeddualstatus": "fspuchangeddualstatus",
    "pcunetstatus": "pcunetstatus",
    "ale_external_link_changed": "ale_external_link_changed",
    "tmsconnectionopen": "tmsconnectionopen",
    "fspbocstatus": "fspbocstatus",
    "fspuchangedmodereply": "fspuchangedmodereply",
    "duallinkstatuschanged": "duallinkstatuschanged",
    "fspuchangedstatus": "fspuchangedstatus",
    "fspuchangedorders": "fspuchangedorders",
    "fspustasindications": "fspustasindications",
    "tmscommandresponse": "tmscommandresponse",
}

CLASS_VALID = {"A", "B", "C", "D"}


def _read_lines_universal(filepath: str, encoding: str) -> list:
    """
    Read all lines from a file, robust to ANY line-ending style:
    \\n (Unix), \\r\\n (Windows), \\r (old Mac), or even mixed/inconsistent
    endings within the same file (common in exports stitched together
    from multiple systems).

    Python's normal text-mode readlines() only reliably splits on \\n
    under universal-newline translation, but some exported files use
    bare \\r with no \\n at all, which silently collapses the whole file
    into a single line if not handled explicitly.
    """
    with open(filepath, "rb") as f:
        raw = f.read()
    text = raw.decode(encoding, errors="replace")
    # Normalise all line-ending variants to \n, then split
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    lines = text.split("\n")
    # Re-add the \n pandas/csv readers don't need it, but keep behaviour
    # consistent with f.readlines() (each line WITHOUT trailing \n is fine
    # since all downstream code already strips \n\r)
    return [line for line in lines]


def _detect_encoding(filepath: str) -> str:
    if HAS_CHARDET:
        with open(filepath, "rb") as f:
            raw = f.read(65536)
        result = chardet.detect(raw)
        return result.get("encoding") or "utf-8"
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            with open(filepath, encoding=enc) as f:
                f.read(1024)
            return enc
        except (UnicodeDecodeError, LookupError):
            pass
    return "latin-1"


def _detect_delimiter(filepath: str, encoding: str, skip: int = 0) -> str:
    """
    Detect the real delimiter by looking at the header row specifically.
    The header row is the most reliable signal because it has NO prose text
    that could contain false-positive delimiters.
    Counts how many fields each candidate delimiter produces in the header
    and picks the one that produces the most fields (≥ 2).
    """
    candidates = [";", ",", "\t", "|"]

    lines = _read_lines_universal(filepath, encoding)
    lines = [l for l in lines if l.strip()]  # drop blank lines for header search
    if len(lines) <= skip:
        return ","
    header = lines[skip].strip()

    if not header:
        return ","

    # Count fields each delimiter produces in the header
    field_counts = {}
    for d in candidates:
        # Split and count non-empty fields
        parts = [p for p in header.split(d) if p.strip()]
        field_counts[d] = len(parts)

    # Pick delimiter that gives the most header fields
    best = max(field_counts, key=field_counts.get)

    # Sanity check: must give at least 2 fields
    if field_counts[best] < 2:
        # Try Python csv sniffer as fallback
        try:
            import csv, io
            sample = header[:2048]
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
            return dialect.delimiter
        except Exception:
            pass
        return ","

    return best


def _split_semicolon_respecting_quotes(s: str) -> list:
    """
    Split a string on ';' while respecting '"..."' quoted segments
    (so semicolons inside quoted text are NOT treated as delimiters).
    The quotes themselves are stripped from the output.
    """
    parts = []
    current = []
    in_quotes = False
    for c in s:
        if c == '"':
            in_quotes = not in_quotes
            continue
        if c == ";" and not in_quotes:
            parts.append("".join(current))
            current = []
            continue
        current.append(c)
    parts.append("".join(current))
    return parts


def _looks_like_padded_quoted_format(lines: list) -> bool:
    """
    Heuristic: returns True if the majority of sampled DATA lines (not the
    header) start with a quote character and end with a long run of
    trailing commas — the signature of the wrapped/padded export format
    where each entire data row is one big quoted CSV field, e.g.:

        "26-05-02;02:58:17;alarm;""Some message"";....",,,,,,,,,,,,,,

    Note: in real SPU exports the HEADER row is typically NOT wrapped in
    quotes (just 'Date;Time;Type;...,,,,,,'), only data rows are — so we
    check rows 1+ rather than row 0.
    """
    sample = [l for l in lines[1:16] if l.strip()]
    if not sample:
        sample = [l for l in lines[:15] if l.strip()]
    if not sample:
        return False
    hits = 0
    for l in sample:
        s = l.rstrip("\n\r")
        if s.startswith('"') and re.search(r",{3,}$", s):
            hits += 1
        elif re.fullmatch(r",+", s):
            hits += 1  # fully blank padded row
    return hits >= len(sample) * 0.5


def _parse_padded_quoted_csv(content_lines_raw: list, log) -> pd.DataFrame:
    """
    Parse the 'entire row wrapped in one quoted CSV field, padded with
    trailing commas' format using Python's csv module to correctly handle
    doubled inner quotes (""..."") and stray literal commas inside the
    quoted text (which would otherwise corrupt naive rstrip-based parsing).

    Format example (one logical row):
        "26-05-02;02:58:17;alarm;""Some, message"";Spu;605;...",,,,,,,,

    Strategy:
      1. Feed each raw line through csv.reader(delimiter=',', quotechar='"')
         — this correctly un-escapes doubled quotes and handles the case
         where a literal comma inside the text breaks the field boundary
         (the real content then spills across row[0], row[1], row[2]...).
      2. Reassemble the real content by re-joining all leading fields with
         ',' up until we hit the trailing all-empty padding.
      3. Split that reassembled string on ';' (respecting any remaining
         quoted segments) to get the final field values.
    """
    import csv as csv_module, io

    def _real_content(row: list) -> str:
        """Re-join fields that got split by a stray comma inside the
        quoted text, stopping at the start of the trailing empty padding."""
        parts = []
        for i, f in enumerate(row):
            if f == "" and all(x == "" for x in row[i:]):
                break
            parts.append(f)
        return ",".join(parts)

    text = "\n".join(content_lines_raw)
    reader = csv_module.reader(io.StringIO(text), delimiter=",", quotechar='"')

    header = None
    rows = []
    n_cols_expected = None
    skipped = 0

    for i, row in enumerate(reader):
        content = _real_content(row)
        if i == 0:
            header = _split_semicolon_respecting_quotes(content)
            n_cols_expected = len(header)
            log(f"  Unwrapped header ({n_cols_expected} cols): {header}")
            continue
        if not content.strip():
            continue
        parts = _split_semicolon_respecting_quotes(content)
        if len(parts) != n_cols_expected:
            # Try to recover: pad short rows, merge overflow into last col
            if len(parts) < n_cols_expected:
                parts += [""] * (n_cols_expected - len(parts))
            else:
                parts = parts[:n_cols_expected - 1] + [
                    ";".join(parts[n_cols_expected - 1:])]
            skipped += 0  # recovered, not skipped
        rows.append(parts)

    if skipped:
        log(f"  {skipped} rows had field-count mismatches (recovered via merge/pad)")

    if not header or n_cols_expected < 2:
        return pd.DataFrame()

    df = pd.DataFrame(rows, columns=header)
    log(f"  Unwrapped parse: {len(df):,} rows × {len(df.columns)} columns")
    return df


def _load_csv_robust(filepath: str, encoding: str, skip: int, log_fn) -> pd.DataFrame:
    """
    Robustly load a CSV regardless of delimiter, quoting, line-ending style,
    or mixed formats.

    Handles every real-world SPU export variant:
      A) Header and data use the same delimiter (normal)
      B) Header is quoted as one string: "Date;Time;..." + data uses ','
      C) Header uses ';' + data uses ',' (true mixed)
      D) File uses \\r-only or mixed line endings (old Mac / re-saved exports)
      E) Entire row wrapped in quotes + doubled inner quotes + padded with
         trailing commas to a fixed column width (Excel "save as CSV" of a
         single text column containing semicolon-delimited data)
      F) Any combination of the above
    """
    def log(m):
        if log_fn: log_fn(m)

    candidates = [";", ",", "\t", "|"]
    import csv as csv_module, io

    # Read all lines using universal newline handling — robust to \r, \n, \r\n
    all_lines = _read_lines_universal(filepath, encoding)

    content_lines_raw = all_lines[skip:]
    log(f"  Lines after universal newline split: {len(all_lines)} total, "
        f"{len(content_lines_raw)} after skipping {skip}")

    if not content_lines_raw:
        log("  No content lines found — returning empty DataFrame")
        return pd.DataFrame()

    # ── Strategy 0: quote-wrapped, padded-comma rows ──────────────────────────
    if _looks_like_padded_quoted_format(content_lines_raw):
        log("  Detected quote-wrapped / comma-padded row format — unwrapping…")
        df = _parse_padded_quoted_csv(content_lines_raw, log)
        if len(df) > 0 and len(df.columns) >= 2:
            return df
        log("  Padded-quote parse produced no rows — falling back to standard parsing")

    # ── Standard path (non-wrapped formats) ───────────────────────────────────
    content_lines = [l for l in content_lines_raw if l.strip() != ""]
    if not content_lines:
        log("  No content lines found — returning empty DataFrame")
        return pd.DataFrame()

    header_raw  = content_lines[0].rstrip("\n\r")
    data_lines  = [l.rstrip("\n\r") for l in content_lines[1:] if l.strip()]

    # ── Step 1: detect header delimiter ──────────────────────────────────────
    # Strip surrounding quotes from header (handles "Date;Time;...")
    header_clean = header_raw.strip().strip('"\'')

    hdr_field_counts = {}
    for d in candidates:
        parts = [p for p in header_clean.split(d) if p.strip()]
        hdr_field_counts[d] = len(parts)
    hdr_delim = max(hdr_field_counts, key=hdr_field_counts.get)
    n_cols = hdr_field_counts[hdr_delim]

    col_names = [c.strip().strip('"\'') for c in header_clean.split(hdr_delim)]
    log(f"Detected delimiter: {repr(hdr_delim)} → {n_cols} columns")

    if n_cols < 2:
        log("  Cannot parse header — returning empty DataFrame")
        return pd.DataFrame()

    # ── Step 2: detect DATA delimiter from first 10 data rows ────────────────
    data_sample = "\n".join(data_lines[:10])
    dat_field_counts = {}
    for d in candidates:
        # Count fields per line, take average
        counts = [len(l.split(d)) for l in data_lines[:10] if l.strip()]
        dat_field_counts[d] = sum(counts) / len(counts) if counts else 0

    # Pick delimiter that gives field count closest to n_cols
    dat_delim = min(candidates,
                    key=lambda d: abs(dat_field_counts[d] - n_cols))

    # ── Step 3: if header and data delimiter agree — use pandas directly ──────
    if hdr_delim == dat_delim:
        import io
        # Rebuild content with clean (unquoted) header
        clean_content = hdr_delim.join(col_names) + "\n" + "\n".join(data_lines)
        try:
            df = pd.read_csv(io.StringIO(clean_content), sep=hdr_delim,
                             dtype=str, engine="python", on_bad_lines="skip")
            if len(df.columns) >= 2:
                return df
        except Exception as e:
            log(f"  pandas read failed ({e}) — falling back to manual parse")

    # ── Step 4: mixed delimiter — parse data rows with dat_delim ─────────────
    log(f"  Mixed delimiters: header={repr(hdr_delim)}, data={repr(dat_delim)}")
    rows = []
    for line in data_lines:
        if not line.strip():
            continue
        parts = line.split(dat_delim)
        if len(parts) < n_cols:
            parts += [""] * (n_cols - len(parts))
        elif len(parts) > n_cols:
            # Merge overflow into last column (comma-in-text protection)
            parts = parts[:n_cols - 1] + [dat_delim.join(parts[n_cols - 1:])]
        rows.append(parts)

    df = pd.DataFrame(rows, columns=col_names)
    return df


def _normalise_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Rename columns using COL_ALIASES, create missing ones as empty."""
    renamed = {}
    for col in df.columns:
        key = re.sub(r"[\s\-_]+", "", str(col)).lower()
        if key in COL_ALIASES:
            renamed[col] = COL_ALIASES[key]
        else:
            # partial match
            for alias, canonical in COL_ALIASES.items():
                if alias in key or key in alias:
                    renamed[col] = canonical
                    break
    df = df.rename(columns=renamed)
    # De-duplicate: keep first occurrence of each canonical col
    seen = set()
    keep = []
    for c in df.columns:
        if c not in seen:
            seen.add(c)
            keep.append(c)
    df = df[keep]
    for col in CLEAN_COLS:
        if col not in df.columns:
            df[col] = pd.NA
    return df[CLEAN_COLS + [c for c in df.columns if c not in CLEAN_COLS]]


def _split_datetime(df: pd.DataFrame) -> pd.DataFrame:
    """
    Robustly split combined datetime values into separate Date and Time columns.
    Handles all common cases:
      - Timestamp/combined column → mapped to Time by aliases → Time contains full datetime
      - Separate Date col contains full datetime string
      - Any unknown column contains a datetime string
    """
    DATE_RE = r"(\d{4}-\d{2}-\d{2})"
    TIME_RE = r"(\d{2}:\d{2}:\d{2})"

    def _looks_like_datetime(series):
        """Return True if series contains combined datetime strings."""
        sample = series.dropna().astype(str)
        if len(sample) == 0:
            return False
        s = sample.iloc[0]
        return bool(re.search(DATE_RE, s) and re.search(TIME_RE, s))

    # Case 1: Time column has full datetime (e.g. '2026-05-02 02:58:17')
    # This happens when 'Timestamp' is aliased to 'Time'
    if "Time" in df.columns and _looks_like_datetime(df["Time"]):
        src = df["Time"].astype(str)
        df["Date"] = src.str.extract(DATE_RE)[0]
        df["Time"] = src.str.extract(TIME_RE)[0]
        return df

    # Case 2: Date column has full datetime
    if "Date" in df.columns and _looks_like_datetime(df["Date"]):
        src = df["Date"].astype(str)
        df["Date"] = src.str.extract(DATE_RE)[0]
        df["Time"] = src.str.extract(TIME_RE)[0]
        return df

    # Case 3: Date is empty/missing — scan ALL columns for a datetime string
    date_missing = ("Date" not in df.columns or
                    df["Date"].isna().all() or
                    df["Date"].astype(str).str.strip().isin(["", "nan", "NaT"]).all())
    if date_missing:
        for col in df.columns:
            if _looks_like_datetime(df[col]):
                src = df[col].astype(str)
                df["Date"] = src.str.extract(DATE_RE)[0]
                df["Time"] = src.str.extract(TIME_RE)[0]
                return df

    # Case 4: Date exists but Time is missing/empty — try splitting Date col
    time_missing = ("Time" not in df.columns or
                    df["Time"].isna().all() or
                    df["Time"].astype(str).str.strip().isin(["", "nan", "NaT"]).all())
    if time_missing and "Date" in df.columns:
        src = df["Date"].astype(str)
        if src.str.contains(TIME_RE).any():
            df["Date"] = src.str.extract(DATE_RE)[0]
            df["Time"] = src.str.extract(TIME_RE)[0]

    return df


def _clean_date(val) -> str:
    if pd.isna(val):
        return pd.NA
    s = str(val).strip()
    # Already good (4-digit year)
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        return s
    # Try common formats — includes 2-digit-year YY-MM-DD (e.g. '26-05-02')
    for fmt in ("%Y-%m-%d", "%y-%m-%d", "%d/%m/%Y", "%d/%m/%y",
                "%d-%m-%Y", "%m/%d/%Y", "%d.%m.%Y", "%d.%m.%y", "%Y%m%d"):
        try:
            return datetime.strptime(s[:10], fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    # Extract from mixed string (4-digit year)
    m = re.search(r"(\d{4}-\d{2}-\d{2})", s)
    if m:
        return m.group(1)
    # Extract 2-digit-year pattern e.g. '26-05-02' → assume 20YY
    m = re.match(r"^(\d{2})-(\d{2})-(\d{2})$", s)
    if m:
        yy, mm, dd = m.groups()
        try:
            return datetime.strptime(f"20{yy}-{mm}-{dd}", "%Y-%m-%d").strftime("%Y-%m-%d")
        except ValueError:
            pass
    return pd.NA


def _clean_time(val) -> str:
    if pd.isna(val):
        return pd.NA
    s = str(val).strip()
    # HH:MM:SS already
    if re.fullmatch(r"\d{2}:\d{2}:\d{2}", s):
        return s
    # HH:MM
    if re.fullmatch(r"\d{2}:\d{2}", s):
        return s + ":00"
    # Extract time portion
    m = re.search(r"(\d{1,2}:\d{2}:\d{2})", s)
    if m:
        parts = m.group(1).split(":")
        return f"{int(parts[0]):02d}:{parts[1]}:{parts[2]}"
    m = re.search(r"(\d{1,2}:\d{2})", s)
    if m:
        parts = m.group(1).split(":")
        return f"{int(parts[0]):02d}:{parts[1]}:00"
    return pd.NA


def _clean_class(val) -> str:
    if pd.isna(val):
        return pd.NA
    s = str(val).strip().upper()
    if s in CLASS_VALID:
        return s
    # Numeric severity → class mapping (common in some SPU exports)
    num_map = {"1": "A", "2": "B", "3": "C", "4": "D",
               "CRITICAL": "A", "MAJOR": "B", "MINOR": "C",
               "INFO": "D", "INFORMATIONAL": "D", "WARNING": "C"}
    return num_map.get(s, pd.NA)


def _clean_code(val):
    if pd.isna(val):
        return pd.NA
    try:
        return float(str(val).strip())
    except ValueError:
        return pd.NA


def _clean_str(val) -> str:
    if pd.isna(val):
        return pd.NA
    s = str(val).strip()
    # Remove common artefacts: leading/trailing quotes, null-like strings
    s = s.strip('"\'')
    if s.lower() in ("nan", "none", "null", "", "-", "n/a", "na"):
        return pd.NA
    return s


def _clean_type(val) -> str:
    if pd.isna(val):
        return pd.NA
    s = str(val).strip().lower().replace(" ", "").replace("-", "").replace("_", "")
    return TYPE_ALIASES.get(s, str(val).strip())


def _drop_junk_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Remove header repetitions, fully empty rows, and separator lines."""
    # Drop rows where Date looks like a column header
    if "Date" in df.columns:
        mask = df["Date"].astype(str).str.lower().isin(
            ["date", "datum", "nan", "", "time", "timestamp"])
        df = df[~mask]
    # Drop fully empty rows
    df = df.dropna(how="all")
    # Drop rows where every value is the same (separator rows)
    def _all_same(row):
        vals = [str(v) for v in row if str(v) not in ("nan", "")]
        return len(set(vals)) == 1 and len(vals) > 2
    mask2 = df.apply(_all_same, axis=1)
    df = df[~mask2]
    return df.reset_index(drop=True)


def clean_raw_file(filepath: str, log_fn=None) -> pd.DataFrame:
    """
    Master cleaning function. Accepts CSV or XLSX, returns clean DataFrame
    with CLEAN_COLS columns.  log_fn(msg) is called with progress messages.
    """
    def log(msg):
        if log_fn:
            log_fn(msg)

    ext = Path(filepath).suffix.lower()
    log(f"Detected file type: {ext.upper()}")

    # ── Load raw ──────────────────────────────────────────────────────────────
    if ext in (".xlsx", ".xls"):
        log("Reading Excel file…")
        # Try to detect header row (skip metadata rows at top)
        raw_peek = pd.read_excel(filepath, header=None, nrows=20)
        header_row = 0
        for i, row in raw_peek.iterrows():
            vals = [str(v).lower().strip() for v in row if str(v) not in ("nan", "")]
            if any(v in ("date", "time", "type", "alarm", "information") for v in vals):
                header_row = i
                break
        df = pd.read_excel(filepath, header=header_row)

    elif ext in (".csv", ".txt", ".log", ".tsv"):
        encoding = _detect_encoding(filepath)
        log(f"Detected encoding: {encoding}")

        # Count comment/metadata lines at top to skip (universal newline safe)
        all_lines_peek = _read_lines_universal(filepath, encoding)
        skip = 0
        for line in all_lines_peek:
            stripped = line.strip()
            if stripped == "":
                continue
            if (stripped.startswith("#") or stripped.startswith("//")
                    or re.match(r"^[=\-\*]{3,}", stripped)):
                skip += 1
            else:
                break
        log(f"Skipping {skip} header/comment lines…")

        # Robust CSV loader — handles any delimiter, mixed formats, quoted
        # headers, and any line-ending style (\\n, \\r\\n, \\r, or mixed)
        df = _load_csv_robust(filepath, encoding, skip, log_fn)
    else:
        # Unknown extension — try robust loader
        encoding = _detect_encoding(filepath)
        df = _load_csv_robust(filepath, encoding, 0, log_fn)

    log(f"Raw shape: {df.shape[0]:,} rows × {df.shape[1]} columns")
    log(f"Raw columns: {df.columns.tolist()[:8]}{'…' if len(df.columns)>8 else ''}")

    # ── Normalise column names ────────────────────────────────────────────────
    log("Normalising column names…")
    df = _normalise_columns(df)

    # ── Split combined datetime ───────────────────────────────────────────────
    df = _split_datetime(df)

    # ── Drop junk rows ────────────────────────────────────────────────────────
    log("Removing junk/header rows…")
    before = len(df)
    df = _drop_junk_rows(df)
    log(f"  Removed {before - len(df)} junk rows")

    # ── Clean each column ─────────────────────────────────────────────────────
    log("Cleaning Date column…")
    df["Date"] = df["Date"].apply(_clean_date)

    log("Cleaning Time column…")
    df["Time"] = df["Time"].apply(_clean_time)

    log("Cleaning Type column…")
    df["Type"] = df["Type"].apply(_clean_type)

    log("Cleaning Class column…")
    df["Class"] = df["Class"].apply(_clean_class)

    log("Cleaning Code column…")
    df["Code"] = df["Code"].apply(_clean_code)

    for col in ["Information", "Group", "Status", "Cause", "Consequence",
                "Action", "Parameters"]:
        df[col] = df[col].apply(_clean_str)

    # ── Drop rows with no Date or Time (completely unparseable) ───────────────
    before = len(df)
    df = df.dropna(subset=["Date", "Time"])
    dropped = before - len(df)
    if dropped:
        log(f"  Dropped {dropped} rows with unparseable date/time")

    # ── Sort ──────────────────────────────────────────────────────────────────
    df = df.sort_values(["Date", "Time"]).reset_index(drop=True)

    log(f"Clean shape: {len(df):,} rows")
    return df[CLEAN_COLS]


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 2 — EXCEL REPORT ENGINE
# ══════════════════════════════════════════════════════════════════════════════

PALETTE = {
    "RED_BG":      "FFCCCC", "ORANGE_BG":   "FFE5CC",
    "YELLOW_BG":   "FFFACD", "BLUE_BG":     "CCE5FF",
    "HEADER_DARK": "1F3864", "HEADER_MID":  "2E75B6",
    "HEADER_LIGHT":"4472C4", "WHITE":        "FFFFFF",
    "LIGHT_GREY":  "F2F2F2", "MID_GREY":    "D9D9D9",
    "RED_TEXT":    "C00000", "RED_DARK":    "C00000",
    "ORANGE_DARK": "C55A11", "ORANGE_TEXT": "E26B0A",
    "DARK_TEXT":   "1F1F1F", "ALT_ROW_A":   "FFE0E0",
    "ALT_ROW_B":   "FFF0DC",
}

CLASS_META = {
    "A": {"label": "🔴 CLASS A – CRITICAL", "severity": "IMMEDIATE ACTION",
          "bg": PALETTE["RED_BG"],    "alt_bg": PALETTE["ALT_ROW_A"],
          "hdr_bg": PALETTE["RED_DARK"],  "txt": PALETTE["RED_TEXT"],
          "desc": "Critical fault – system safety/availability at risk"},
    "B": {"label": "🟠 CLASS B – MAJOR",    "severity": "PROMPT ACTION",
          "bg": PALETTE["ORANGE_BG"], "alt_bg": PALETTE["ALT_ROW_B"],
          "hdr_bg": PALETTE["ORANGE_DARK"], "txt": PALETTE["ORANGE_TEXT"],
          "desc": "Major fault – significant degradation of system"},
    "C": {"label": "🟡 CLASS C – MINOR",    "severity": "SCHEDULED ACTION",
          "bg": PALETTE["YELLOW_BG"], "alt_bg": "FFFFF0",
          "hdr_bg": "806000", "txt": "806000",
          "desc": "Minor fault – monitor and plan maintenance"},
    "D": {"label": "🔵 CLASS D – INFO",     "severity": "NO ACTION",
          "bg": PALETTE["BLUE_BG"],   "alt_bg": "E8F4FF",
          "hdr_bg": "1F497D", "txt": "1F497D",
          "desc": "Informational event – no action required"},
}


def _fill(h):  return PatternFill("solid", start_color=h, fgColor=h)
def _bdr(s="thin"):
    e = Side(style=s)
    return Border(left=e, right=e, top=e, bottom=e)
def _fnt(sz=9, bold=False, color=PALETTE["DARK_TEXT"], name="Arial"):
    return Font(name=name, size=sz, bold=bold, color=color)
def _al(h="left", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

def _set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _banner(ws, row, text, bg, fg="FFFFFF", size=12, span=11):
    ws.merge_cells(f"A{row}:{get_column_letter(span)}{row}")
    c = ws[f"A{row}"]
    c.value, c.font = text, Font(name="Arial", size=size, bold=True, color=fg)
    c.fill, c.alignment = _fill(bg), _al("center")
    ws.row_dimensions[row].height = 26
    return row + 1

def _section(ws, row, text, span=11):
    ws.merge_cells(f"A{row}:{get_column_letter(span)}{row}")
    c = ws[f"A{row}"]
    c.value, c.font = text, Font(name="Arial", size=10, bold=True, color="FFFFFF")
    c.fill, c.alignment = _fill(PALETTE["HEADER_MID"]), _al("left", indent=1)
    ws.row_dimensions[row].height = 20
    return row + 1

def _kv(ws, row, label, value, span=5):
    ws[f"A{row}"] = label
    ws[f"A{row}"].font  = _fnt(9, bold=True)
    ws[f"A{row}"].fill  = _fill(PALETTE["MID_GREY"])
    ws[f"A{row}"].alignment = _al("left", indent=1)
    ws[f"A{row}"].border = _bdr()
    ws.merge_cells(f"B{row}:{get_column_letter(1+span)}{row}")
    c = ws[f"B{row}"]
    c.value, c.font = value, _fnt(9)
    c.fill, c.alignment, c.border = _fill(PALETTE["WHITE"]), _al("left", indent=1), _bdr()
    ws.row_dimensions[row].height = 16
    return row + 1

def _hdr_row(ws, row, cols, bg, fg="FFFFFF", h=22):
    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=col)
        c.font, c.fill = Font(name="Arial", size=9, bold=True, color=fg), _fill(bg)
        c.alignment, c.border = _al("center", wrap=True), _bdr()
    ws.row_dimensions[row].height = h
    return row + 1

def _sheet_hyperlink(ws_src, row, col, target_ws, label):
    safe = target_ws.title.replace("'", "''")
    c = ws_src.cell(row=row, column=col, value=label)
    c.hyperlink = f"#'{safe}'!A1"
    c.font = Font(name="Arial", size=8, bold=True, color="1F497D", underline="single")
    c.fill, c.alignment, c.border = _fill("EEF4FF"), _al("center"), _bdr()


def _normalise_info(text: str) -> str:
    if not isinstance(text, str):
        return text
    text = re.sub(r"([0-9a-fA-F]{2}:){5}[0-9a-fA-F]{2}", "(mac address)", text)
    text = re.sub(r"\b\d{1,3}(?:\.\d{1,3}){3}\b", "(ip address)", text)
    text = re.sub(r"\b0x[0-9a-fA-F]+\b", "(id)", text)
    text = re.sub(r"\b([A-Z]{1,5})\d{4,}\b", r"\1(id)", text)
    text = re.sub(r"\b\d{4,}\b", "(id)", text)
    text = re.sub(r"  +", " ", text).strip()
    return text


# ── Sheet builders ────────────────────────────────────────────────────────────

def _sheet_summary(wb, filtered, alarms, start_t, end_t):
    ws = wb.active
    ws.title = "📊 Summary"
    _set_widths(ws, [38, 10, 10, 52, 16, 12])

    alarm_count = len(alarms)
    date_label  = str(filtered["Date"].iloc[0]) if len(filtered) else "N/A"
    sheet_map   = {s.title: s for s in wb.worksheets}

    r = 1
    r = _banner(ws, r, "SPU LOG FAULT ANALYSIS REPORT", PALETTE["HEADER_DARK"], size=14, span=6)
    r = _banner(ws, r, f"Period: {start_t[:5]} – {end_t[:5]}   |   Date: {date_label}   |   "
                f"Total Entries: {len(filtered):,}   |   Alarms: {alarm_count:,}",
                PALETTE["HEADER_MID"], size=9, span=6)
    ws.row_dimensions[r-1].height = 18
    r += 1

    # ── Overview ──────────────────────────────────────────────────────────────
    r = _section(ws, r, "📋  LOG OVERVIEW", span=6)
    r = _kv(ws, r, "Analysis Period",   f"{start_t[:5]} to {end_t[:5]}  ({date_label})")
    r = _kv(ws, r, "Total Log Entries", f"{len(filtered):,}")
    r = _kv(ws, r, "Total Alarm Events",f"{alarm_count:,}")
    r = _kv(ws, r, "Non-Alarm Events",  f"{len(filtered) - alarm_count:,}")
    r += 1

    # ── Fault classification ──────────────────────────────────────────────────
    r = _section(ws, r, "🚨  FAULT CLASSIFICATION BREAKDOWN", span=6)
    r = _hdr_row(ws, r, ["Fault Class","Class Label","Count","% of Alarms","Severity","Description"],
                 PALETTE["HEADER_LIGHT"])
    for cls, meta in CLASS_META.items():
        cnt = len(alarms[alarms["Class"] == cls]) if "Class" in alarms.columns else 0
        pct = f"{cnt/alarm_count*100:.1f}%" if alarm_count else "0%"
        for ci, val in enumerate([f"Class {cls}", meta["label"], cnt, pct,
                                   meta["severity"], meta["desc"]], 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.fill  = _fill(meta["bg"])
            c.font  = Font(name="Arial", size=9, bold=(ci<=2),
                           color=meta["txt"] if ci==1 else PALETTE["DARK_TEXT"])
            c.alignment = _al("center" if ci in [1,3,4,5] else "left",
                              indent=0 if ci in [1,3,4,5] else 1)
            c.border = _bdr()
        ws.row_dimensions[r].height = 17
        r += 1
    r += 1

    # ── Event type distribution + alarm tree ─────────────────────────────────
    r = _section(ws, r, "⚡  EVENT TYPE DISTRIBUTION & ALARM DRILL-DOWN", span=6)
    r = _hdr_row(ws, r, ["Event Type / Alarm Sub-category", "Count", "% of Total",
                          "Suggestion / Action", "→ Sheet"], PALETTE["HEADER_DARK"])

    total = len(filtered)
    type_counts = filtered["Type"].value_counts()

    # Pre-build alarm tree with normalised descriptions
    alarm_tree = {}
    if "Class" in alarms.columns and "Information" in alarms.columns:
        for cls in ["A","B","C","D"]:
            sub = alarms[alarms["Class"]==cls].copy()
            sub["_norm"] = sub["Information"].apply(
                lambda v: _normalise_info(str(v)) if pd.notna(v) else "(no description)")
            info_data = {}
            for norm_key, grp in sub.groupby("_norm", dropna=False):
                disp   = norm_key if len(norm_key)<=72 else norm_key[:69]+"…"
                action = ""
                if "Action" in grp.columns:
                    acts = grp["Action"].dropna().unique()
                    if len(acts):
                        a = str(acts[0])
                        action = a if len(a)<=80 else a[:77]+"…"
                info_data[disp] = {"count": len(grp), "action": action}
            alarm_tree[cls] = dict(sorted(info_data.items(), key=lambda x: -x[1]["count"]))

    class_sheet_map = {
        "A": "🔴 Class A – Critical", "B": "🟠 Class B – Major",
        "C": "🟡 Class C – Minor",    "D": "🔵 Class D – Info",
    }

    for i, (etype, cnt) in enumerate(type_counts.items()):
        pct      = f"{cnt/total*100:.1f}%"
        is_alarm = etype == "alarm"
        bg       = PALETTE["HEADER_MID"] if is_alarm else (
                   PALETTE["LIGHT_GREY"] if i%2==0 else PALETTE["WHITE"])
        fg       = "FFFFFF" if is_alarm else PALETTE["DARK_TEXT"]

        for ci, val in enumerate([("▶  " if is_alarm else "   ")+etype.upper(),
                                   cnt, pct,
                                   "See alarm drill-down ↓" if is_alarm
                                   else "Informational – no action required"], 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.font = Font(name="Arial", size=9, bold=is_alarm, color=fg)
            c.fill, c.border = _fill(bg), _bdr()
            c.alignment = _al("center" if ci in [2,3] else "left", indent=1 if ci in [1,4] else 0)

        # Hyperlink to dedicated sheet
        if is_alarm and "⚠️ All Alarms" in sheet_map:
            _sheet_hyperlink(ws, r, 5, sheet_map["⚠️ All Alarms"], "→ All Alarms")
        elif not is_alarm:
            # Link to dedicated event-type sheet
            sheet_title = f"📋 {re.sub(r'[\\/*?:\\[\\]]', '_', etype)[:28]}"
            if sheet_title in sheet_map:
                _sheet_hyperlink(ws, r, 5, sheet_map[sheet_title], "→ Details")
            else:
                ws.cell(row=r, column=5).fill  = _fill(bg)
                ws.cell(row=r, column=5).border = _bdr()
        else:
            ws.cell(row=r, column=5).fill  = _fill(bg)
            ws.cell(row=r, column=5).border = _bdr()
        ws.row_dimensions[r].height = 17
        r += 1

        # ── Alarm tree expansion ──────────────────────────────────────────────
        if is_alarm:
            for cls, meta in CLASS_META.items():
                cls_cnt = len(alarms[alarms["Class"]==cls]) if "Class" in alarms.columns else 0
                cls_pct = f"{cls_cnt/total*100:.1f}%" if total else "0%"
                cls_bg  = meta["bg"]

                c = ws.cell(row=r, column=1,
                            value=f"    ├─ {meta['label']}  ({cls_cnt} events)")
                c.font  = Font(name="Arial", size=9, bold=cls in ["A","B"],
                               color=meta["txt"])
                c.fill, c.border = _fill(cls_bg), _bdr()
                c.alignment = _al("left", indent=2)

                for ci, val in enumerate([cls_cnt, cls_pct, meta["desc"]], 2):
                    cell = ws.cell(row=r, column=ci, value=val)
                    cell.font = _fnt(9, bold=(ci==2 and cls in ["A","B"]))
                    cell.fill, cell.border = _fill(cls_bg), _bdr()
                    cell.alignment = _al("center" if ci in [2,3] else "left",
                                        wrap=(ci==4), indent=1 if ci==4 else 0)

                sht = class_sheet_map.get(cls)
                if sht and sht in sheet_map:
                    _sheet_hyperlink(ws, r, 5, sheet_map[sht], f"→ Class {cls}")
                else:
                    ws.cell(row=r, column=5).fill  = _fill(cls_bg)
                    ws.cell(row=r, column=5).border = _bdr()
                ws.row_dimensions[r].height = 16
                r += 1

                # Sub-event rows
                sub_events = alarm_tree.get(cls, {})
                items = list(sub_events.items())
                for si, (info_text, info_data) in enumerate(items):
                    sub_bg  = {"A":"FFF8F8","B":"FFF5EE","C":"FFFFF8","D":"F0F8FF"}[cls]
                    branch  = "    │    └─ " if si==len(items)-1 else "    │    ├─ "
                    action  = info_data["action"] or meta["severity"]

                    c = ws.cell(row=r, column=1, value=branch+info_text)
                    c.font = _fnt(8)
                    c.fill, c.border = _fill(sub_bg), _bdr()
                    c.alignment = _al("left", wrap=True, indent=1)

                    sub_pct = f"{info_data['count']/alarm_count*100:.1f}%" if alarm_count else "0%"
                    for ci, val in enumerate([info_data["count"], sub_pct, action], 2):
                        cell = ws.cell(row=r, column=ci, value=val)
                        cell.font = Font(name="Arial", size=8,
                                         bold=(ci==4 and cls in ["A","B"]),
                                         color=meta["txt"] if ci==4 and cls in ["A","B"]
                                               else PALETTE["DARK_TEXT"])
                        cell.fill, cell.border = _fill(sub_bg), _bdr()
                        cell.alignment = _al("center" if ci in [2,3] else "left",
                                             wrap=(ci==4), indent=1 if ci==4 else 0)

                    ws.cell(row=r, column=5).fill  = _fill(sub_bg)
                    ws.cell(row=r, column=5).border = _bdr()
                    ws.row_dimensions[r].height = max(15, min(40, len(info_text)//4))
                    r += 1

    ws.freeze_panes = "A3"


def _sheet_class(wb, alarms, cls, start_t, end_t):
    meta   = CLASS_META[cls]
    faults = alarms[alarms["Class"]==cls].copy() if "Class" in alarms.columns else pd.DataFrame()
    names  = {"A":"Critical","B":"Major","C":"Minor","D":"Info"}
    emojis = {"A":"🔴","B":"🟠","C":"🟡","D":"🔵"}
    ws     = wb.create_sheet(f"{emojis[cls]} Class {cls} – {names[cls]}")
    _set_widths(ws, [6,10,10,22,55,10,8,12,48,55,50])

    r = 1
    r = _banner(ws, r, f"{meta['label']} — {meta['severity']}", meta["hdr_bg"], span=11)
    ws.merge_cells("A2:K2")
    ws["A2"] = (f"Period: {start_t[:5]}–{end_t[:5]}   |   "
                f"Total {emojis[cls]} Class {cls}: {len(faults)}   |   {meta['severity']}")
    ws["A2"].font = Font(name="Arial", size=9, bold=True, color=meta["txt"])
    ws["A2"].fill = _fill(meta["bg"])
    ws["A2"].alignment = _al("center")
    ws.row_dimensions[2].height = 18

    cols = ["#","Date","Time","Type","Description / Information",
            "Group","Code","Status","Cause","Consequence","Action to Take"]
    _hdr_row(ws, 3, cols, meta["hdr_bg"])

    for idx, (_, row) in enumerate(faults.iterrows(), 1):
        bg = _fill(meta["bg"] if idx%2==1 else meta["alt_bg"])
        rd = [idx, row.get("Date"), row.get("Time"), row.get("Type"),
              row.get("Information"), row.get("Group"), row.get("Code"),
              row.get("Status"), row.get("Cause"), row.get("Consequence"), row.get("Action")]
        for ci, val in enumerate(rd, 1):
            c = ws.cell(row=3+idx, column=ci, value=val if pd.notna(val) else "")
            c.font = _fnt(8, bold=(ci==9))
            c.fill, c.border = bg, _bdr()
            c.alignment = _al("center" if ci in [1,3,6,7,8] else "left",
                              wrap=(ci in [5,9,10,11]))
        ws.row_dimensions[3+idx].height = 15

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(cols))}3"


def _sheet_all_alarms(wb, alarms, start_t, end_t):
    ws = wb.create_sheet("⚠️ All Alarms")
    _set_widths(ws, [6,10,10,22,55,10,8,20,10,48,55,50])

    r = 1
    r = _banner(ws, r, "ALL ALARM EVENTS — COLOUR-CODED BY CLASS",
                PALETTE["HEADER_DARK"], span=12)
    ws.merge_cells("A2:L2")
    ws["A2"] = (f"Period: {start_t[:5]}–{end_t[:5]}   |   Total Alarms: {len(alarms)}   |   "
                "🔴 A=Critical  🟠 B=Major  🟡 C=Minor  🔵 D=Info")
    ws["A2"].font = Font(name="Arial", size=9, color="FFFFFF")
    ws["A2"].fill = _fill(PALETTE["HEADER_MID"])
    ws["A2"].alignment = _al("center")
    ws.row_dimensions[2].height = 18

    cols = ["#","Date","Time","Type","Description / Information",
            "Group","Code","Fault Class","Status","Cause","Consequence","Action to Take"]
    _hdr_row(ws, 3, cols, PALETTE["HEADER_DARK"])
    r = 4

    for idx, (_, row) in enumerate(alarms.sort_values("Time").iterrows(), 1):
        cls  = str(row.get("Class","")).strip() if pd.notna(row.get("Class","")) else ""
        meta = CLASS_META.get(cls, {})
        bg   = _fill(meta["bg"]) if meta else _fill(PALETTE["LIGHT_GREY"])
        label = meta.get("label", f"Class {cls}")
        rd = [idx, row.get("Date"), row.get("Time"), row.get("Type"),
              row.get("Information"), row.get("Group"), row.get("Code"),
              label, row.get("Status"), row.get("Cause"),
              row.get("Consequence"), row.get("Action")]
        for ci, val in enumerate(rd, 1):
            c = ws.cell(row=r, column=ci, value=val if pd.notna(val) else "")
            c.font = Font(name="Arial", size=8,
                          bold=(cls in ["A","B"] and ci in [8,9,10]),
                          color=meta.get("txt", PALETTE["DARK_TEXT"]) if ci==8 and meta
                                else PALETTE["DARK_TEXT"])
            c.fill, c.border = bg, _bdr()
            c.alignment = _al("center" if ci in [1,3,6,7] else "left",
                              wrap=(ci in [5,10,11,12]))
        ws.row_dimensions[r].height = 15
        r += 1

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(cols))}3"


def _sheet_fault_ref(wb, alarms, start_t, end_t):
    ws = wb.create_sheet("📖 Fault Code Reference")
    _set_widths(ws, [8,22,22,20,18,48,55,50])

    r = 1
    r = _banner(ws, r, "UNIQUE FAULT CODE REFERENCE — ALL CODES IN THIS PERIOD",
                PALETTE["HEADER_DARK"], size=11, span=8)
    ws.merge_cells("A2:H2")
    ws["A2"] = "Unique alarm codes with classification, cause and recommended action"
    ws["A2"].font = Font(name="Arial", size=9, color="FFFFFF")
    ws["A2"].fill = _fill(PALETTE["HEADER_MID"])
    ws["A2"].alignment = _al("center")
    ws.row_dimensions[2].height = 18

    cols = ["Code","Group","Type","Fault Class","Class Label","Cause","Consequence","Action"]
    _hdr_row(ws, 3, cols, PALETTE["HEADER_LIGHT"])
    r = 4

    gcols = [c for c in ["Code","Group","Type","Class","Cause","Consequence","Action"]
             if c in alarms.columns]
    if len(gcols) >= 3:
        unique = (alarms[alarms["Code"].notna()].groupby(gcols, dropna=False)
                  .size().reset_index(name="_n")
                  .sort_values(["Class","Code"] if "Class" in gcols else ["Code"]))
        for idx, (_, row) in enumerate(unique.iterrows()):
            cls  = str(row.get("Class","")).strip()
            meta = CLASS_META.get(cls, {})
            bg   = _fill(meta["bg"]) if meta else _fill(
                   PALETTE["LIGHT_GREY"] if idx%2==0 else PALETTE["WHITE"])
            rd   = [row.get("Code"), row.get("Group"), row.get("Type"),
                    f"Class {cls}", meta.get("label",""),
                    row.get("Cause"), row.get("Consequence"), row.get("Action")]
            for ci, val in enumerate(rd, 1):
                c = ws.cell(row=r, column=ci, value=val if pd.notna(val) else "")
                c.font = Font(name="Arial", size=8,
                              bold=(ci in [4,5] and cls in ["A","B"]),
                              color=meta.get("txt", PALETTE["DARK_TEXT"]) if ci==5 and meta
                                    else PALETTE["DARK_TEXT"])
                c.fill, c.border = bg, _bdr()
                c.alignment = _al("center" if ci==4 else "left", wrap=(ci in [6,7,8]))
            ws.row_dimensions[r].height = 15
            r += 1

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(cols))}3"


def _sheet_event_type(wb, filtered, etype, start_t, end_t):
    """
    Build a dedicated sheet for a single non-alarm event type,
    showing a sub-event drill-down (normalised Information strings)
    and the full event rows — mirroring the alarm class sheet style.
    """
    rows     = filtered[filtered["Type"] == etype].copy()
    safe     = re.sub(r"[\\/*?:\[\]]", "_", etype)   # Excel tab name safe chars
    title    = safe[:28]                               # max 31 chars for sheet name
    ws       = wb.create_sheet(f"📋 {title}")

    # Pick a neutral colour scheme per event type (cycle through a palette)
    TYPE_COLOURS = [
        {"bg": "E8F4F8", "hdr": "17608A", "txt": "0D3B54"},
        {"bg": "EAF2EA", "hdr": "2D6A2D", "txt": "1A3F1A"},
        {"bg": "F5EDF8", "hdr": "6A3080", "txt": "3D1A4A"},
        {"bg": "FFF3E0", "hdr": "8A5200", "txt": "4A2D00"},
        {"bg": "F0F0F0", "hdr": "404040", "txt": "202020"},
        {"bg": "E0EFFF", "hdr": "1A4A8A", "txt": "0D2A4A"},
        {"bg": "FFF0F0", "hdr": "8A2020", "txt": "4A1010"},
        {"bg": "FFFFF0", "hdr": "7A7A00", "txt": "404000"},
    ]
    all_types  = list(wb._sheets_by_title.keys()) if hasattr(wb, "_sheets_by_title") else []
    colour_idx = hash(etype) % len(TYPE_COLOURS)
    col        = TYPE_COLOURS[colour_idx]
    alt_bg     = col["bg"]
    hdr_bg     = col["hdr"]

    _set_widths(ws, [6, 10, 10, 55, 10, 8, 12, 48, 55, 50])

    r = 1
    r = _banner(ws, r,
                f"EVENT TYPE: {etype.upper()}  —  {len(rows):,} events in period",
                hdr_bg, size=12, span=10)
    ws.merge_cells("A2:J2")
    ws["A2"] = (f"Period: {start_t[:5]}–{end_t[:5]}   |   "
                f"Total events of this type: {len(rows):,}")
    ws["A2"].font = Font(name="Arial", size=9, bold=True, color=col["txt"])
    ws["A2"].fill = _fill(alt_bg)
    ws["A2"].alignment = _al("center")
    ws.row_dimensions[2].height = 18

    # ── Sub-event drill-down (normalised Information) ─────────────────────────
    if "Information" in rows.columns and rows["Information"].notna().any():
        r = _section(ws, r, "🔍  EVENT SUB-TYPE BREAKDOWN", span=10)
        r = _hdr_row(ws, r,
                     ["Sub-category (normalised)", "Count", "% of Type",
                      "Sample Action / Info"], hdr_bg)

        rows["_norm"] = rows["Information"].apply(
            lambda v: _normalise_info(str(v)) if pd.notna(v) else "(no description)")
        grouped = (rows.groupby("_norm", dropna=False)
                       .size()
                       .reset_index(name="cnt")
                       .sort_values("cnt", ascending=False))

        for gi, (_, grow) in enumerate(grouped.iterrows()):
            norm_txt = grow["_norm"]
            cnt      = grow["cnt"]
            pct      = f"{cnt/len(rows)*100:.1f}%" if len(rows) else "0%"
            disp     = norm_txt if len(norm_txt) <= 72 else norm_txt[:69] + "…"

            # Get a sample action from these rows
            sample_rows = rows[rows["_norm"] == norm_txt]
            sample_action = ""
            if "Action" in sample_rows.columns:
                acts = sample_rows["Action"].dropna().unique()
                if len(acts):
                    a = str(acts[0])
                    sample_action = a if len(a) <= 80 else a[:77] + "…"

            sub_bg = alt_bg if gi % 2 == 0 else PALETTE["WHITE"]
            for ci, val in enumerate([disp, cnt, pct, sample_action], 1):
                cell = ws.cell(row=r, column=ci, value=val)
                cell.font = _fnt(8)
                cell.fill, cell.border = _fill(sub_bg), _bdr()
                cell.alignment = _al("center" if ci in [2, 3] else "left",
                                     wrap=(ci in [1, 4]), indent=1 if ci in [1, 4] else 0)
            ws.row_dimensions[r].height = 15
            r += 1

        rows.drop(columns=["_norm"], inplace=True)
        r += 1

    # ── Full event rows ────────────────────────────────────────────────────────
    r = _section(ws, r, "📋  ALL EVENTS OF THIS TYPE", span=10)
    cols = ["#", "Date", "Time", "Information", "Group", "Code",
            "Status", "Cause", "Consequence", "Action"]
    _hdr_row(ws, r, cols, hdr_bg)
    r += 1

    for idx, (_, row) in enumerate(rows.iterrows(), 1):
        bg = _fill(alt_bg if idx % 2 == 1 else PALETTE["WHITE"])
        rd = [idx, row.get("Date"), row.get("Time"), row.get("Information"),
              row.get("Group"), row.get("Code"), row.get("Status"),
              row.get("Cause"), row.get("Consequence"), row.get("Action")]
        for ci, val in enumerate(rd, 1):
            c = ws.cell(row=r, column=ci, value=val if pd.notna(val) else "")
            c.font = _fnt(8)
            c.fill, c.border = bg, _bdr()
            c.alignment = _al("center" if ci in [1, 3, 5, 6, 7] else "left",
                              wrap=(ci in [4, 8, 9, 10]))
        ws.row_dimensions[r].height = 15
        r += 1

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(cols))}3"


def _sheet_event_type_summary(wb, filtered, start_t, end_t):
    """
    A dedicated overview sheet for ALL non-alarm event types,
    showing a tree with sub-category breakdown — mirrors the alarm tree.
    """
    ws = wb.create_sheet("📋 Event Types")
    _set_widths(ws, [42, 10, 10, 52, 16])

    total = len(filtered)
    date_label = str(filtered["Date"].iloc[0]) if len(filtered) else "N/A"
    sheet_map  = {s.title: s for s in wb.worksheets}

    r = 1
    r = _banner(ws, r, "NON-ALARM EVENT TYPE BREAKDOWN — ALL TYPES",
                PALETTE["HEADER_DARK"], size=12, span=5)
    ws.merge_cells("A2:E2")
    ws["A2"] = (f"Period: {start_t[:5]}–{end_t[:5]}   |   Date: {date_label}   |   "
                f"Total Non-Alarm Events: {len(filtered[filtered['Type'] != 'alarm']):,}")
    ws["A2"].font = Font(name="Arial", size=9, color="FFFFFF")
    ws["A2"].fill = _fill(PALETTE["HEADER_MID"])
    ws["A2"].alignment = _al("center")
    ws.row_dimensions[2].height = 18
    r += 1

    r = _hdr_row(ws, r,
                 ["Event Type / Sub-category", "Count", "% of Total",
                  "Description / Action", "→ Sheet"], PALETTE["HEADER_DARK"])

    TYPE_COLOURS = [
        {"bg": "E8F4F8", "hdr": "17608A"},
        {"bg": "EAF2EA", "hdr": "2D6A2D"},
        {"bg": "F5EDF8", "hdr": "6A3080"},
        {"bg": "FFF3E0", "hdr": "8A5200"},
        {"bg": "F0F0F0", "hdr": "404040"},
        {"bg": "E0EFFF", "hdr": "1A4A8A"},
        {"bg": "FFF0F0", "hdr": "8A2020"},
        {"bg": "FFFFF0", "hdr": "7A7A00"},
    ]

    non_alarm = filtered[filtered["Type"] != "alarm"]
    type_counts = non_alarm["Type"].value_counts()

    for ti, (etype, cnt) in enumerate(type_counts.items()):
        col    = TYPE_COLOURS[hash(etype) % len(TYPE_COLOURS)]
        bg     = col["bg"]
        hdr_c  = col["hdr"]
        pct    = f"{cnt/total*100:.1f}%"

        # ── Event type header row ─────────────────────────────────────────────
        c = ws.cell(row=r, column=1, value=f"▶  {etype.upper()}")
        c.font  = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        c.fill, c.border = _fill(hdr_c), _bdr()
        c.alignment = _al("left", indent=1)

        for ci, val in enumerate([cnt, pct, "See sub-categories ↓"], 2):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
            cell.fill, cell.border = _fill(hdr_c), _bdr()
            cell.alignment = _al("center" if ci in [2, 3] else "left",
                                 indent=1 if ci == 4 else 0)

        # Hyperlink to dedicated sheet
        sheet_title = f"📋 {re.sub(r'[\\/*?:\\[\\]]', '_', etype)[:28]}"
        if sheet_title in sheet_map:
            _sheet_hyperlink(ws, r, 5, sheet_map[sheet_title], "→ Details")
        else:
            ws.cell(row=r, column=5).fill  = _fill(hdr_c)
            ws.cell(row=r, column=5).border = _bdr()
        ws.row_dimensions[r].height = 18
        r += 1

        # ── Sub-event rows ────────────────────────────────────────────────────
        type_rows = non_alarm[non_alarm["Type"] == etype].copy()
        if "Information" in type_rows.columns and type_rows["Information"].notna().any():
            type_rows["_norm"] = type_rows["Information"].apply(
                lambda v: _normalise_info(str(v)) if pd.notna(v) else "(no description)")
            grouped = (type_rows.groupby("_norm", dropna=False)
                                .size()
                                .reset_index(name="cnt")
                                .sort_values("cnt", ascending=False))
            items = grouped.iterrows()
            total_sub = len(grouped)

            for si, (_, grow) in enumerate(grouped.iterrows()):
                norm_txt = grow["_norm"]
                sub_cnt  = grow["cnt"]
                sub_pct  = f"{sub_cnt/cnt*100:.1f}%"
                disp     = norm_txt if len(norm_txt) <= 70 else norm_txt[:67] + "…"
                branch   = "    └─ " if si == total_sub - 1 else "    ├─ "

                sample_rows = type_rows[type_rows["_norm"] == norm_txt]
                action = ""
                if "Action" in sample_rows.columns:
                    acts = sample_rows["Action"].dropna().unique()
                    if len(acts):
                        a = str(acts[0])
                        action = a if len(a) <= 80 else a[:77] + "…"

                sub_bg = bg if si % 2 == 0 else PALETTE["WHITE"]
                c = ws.cell(row=r, column=1, value=branch + disp)
                c.font = _fnt(8)
                c.fill, c.border = _fill(sub_bg), _bdr()
                c.alignment = _al("left", wrap=True, indent=1)

                for ci, val in enumerate([sub_cnt, sub_pct, action], 2):
                    cell = ws.cell(row=r, column=ci, value=val)
                    cell.font = _fnt(8)
                    cell.fill, cell.border = _fill(sub_bg), _bdr()
                    cell.alignment = _al("center" if ci in [2, 3] else "left",
                                         wrap=(ci == 4), indent=1 if ci == 4 else 0)

                ws.cell(row=r, column=5).fill  = _fill(sub_bg)
                ws.cell(row=r, column=5).border = _bdr()
                ws.row_dimensions[r].height = max(15, min(38, len(disp) // 4))
                r += 1

    ws.freeze_panes = "A4"


def _sheet_filtered_export(filtered, alarms, start_t, end_t, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Filtered Log"

    ca = len(alarms[alarms["Class"]=="A"]) if "Class" in alarms.columns else 0
    cb = len(alarms[alarms["Class"]=="B"]) if "Class" in alarms.columns else 0
    cc = len(alarms[alarms["Class"]=="C"]) if "Class" in alarms.columns else 0
    cd = len(alarms[alarms["Class"]=="D"]) if "Class" in alarms.columns else 0
    date_lbl = str(filtered["Date"].iloc[0]) if len(filtered) else "N/A"

    _banner(ws, 1,
        f"SPU Log — Filtered: {start_t[:5]} to {end_t[:5]}  |  Date: {date_lbl}  |  "
        f"Total Entries: {len(filtered):,}",
        PALETTE["HEADER_DARK"], size=11, span=12)
    _banner(ws, 2,
        f"Alarms: {len(alarms):,}   |   🔴 A: {ca}   |   🟠 B: {cb}   "
        f"|   🟡 C: {cc}   |   🔵 D: {cd}",
        PALETTE["HEADER_MID"], size=9, span=12)

    cols   = ["#","Date","Time","Type","Information","Group","Code",
              "Status","Class","Cause","Consequence","Action"]
    widths = [5,12,10,22,55,10,7,10,18,42,50,45]
    _set_widths(ws, widths)
    _hdr_row(ws, 3, cols, PALETTE["HEADER_LIGHT"])

    for ri, (_, row) in enumerate(filtered.iterrows(), 4):
        cls = str(row.get("Class","")).strip() if pd.notna(row.get("Class","")) else ""
        bg  = (_fill(CLASS_META[cls]["bg"]) if cls in CLASS_META else
               _fill(PALETTE["LIGHT_GREY"] if ri%2==0 else PALETTE["WHITE"]))
        rd  = [ri-3, row.get("Date"), row.get("Time"), row.get("Type"),
               row.get("Information"), row.get("Group"), row.get("Code"),
               row.get("Status"), row.get("Class"), row.get("Cause"),
               row.get("Consequence"), row.get("Action")]
        for ci, val in enumerate(rd, 1):
            c = ws.cell(row=ri, column=ci, value=val if pd.notna(val) else "")
            c.font = _fnt(8, bold=(cls in ["A","B"]))
            c.fill, c.border = bg, _bdr()
            c.alignment = _al("center" if ci in [1,3,7,8] else "left",
                              wrap=(ci in [5,10,11,12]))
        ws.row_dimensions[ri].height = 15

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(cols))}3"
    wb.save(out_path)


def build_analysis_report(filtered, alarms, start_t, end_t, out_path):
    wb = openpyxl.Workbook()

    # ── 1. Build all alarm detail sheets first ────────────────────────────────
    _sheet_all_alarms(wb, alarms, start_t, end_t)
    for cls in ["A", "B", "C", "D"]:
        _sheet_class(wb, alarms, cls, start_t, end_t)
    _sheet_fault_ref(wb, alarms, start_t, end_t)

    # ── 2. Build a dedicated sheet for every non-alarm event type ─────────────
    non_alarm_types = (filtered[filtered["Type"] != "alarm"]["Type"]
                       .value_counts().index.tolist())
    for etype in non_alarm_types:
        _sheet_event_type(wb, filtered, etype, start_t, end_t)

    # ── 3. Event-type overview sheet (tree for all non-alarm types) ───────────
    _sheet_event_type_summary(wb, filtered, start_t, end_t)

    # ── 4. Summary last — it hyperlinks to everything above ───────────────────
    _sheet_summary(wb, filtered, alarms, start_t, end_t)
    wb.move_sheet("📊 Summary", offset=-(len(wb.worksheets) - 1))

    wb.save(out_path)


def load_and_filter(filepath, start_t, end_t, log_fn=None):
    """Load, clean, and filter by time range. Works for raw CSV or XLSX."""
    ext = Path(filepath).suffix.lower()

    if ext in (".csv", ".txt", ".log", ".tsv"):
        # Raw text file — always run through full cleaning pipeline
        if log_fn: log_fn("Raw CSV detected — running cleaning pipeline…")
        df = clean_raw_file(filepath, log_fn=log_fn)

    else:
        # Excel file — always clean (handles both raw and pre-cleaned XLSX)
        if log_fn: log_fn("Excel file detected — cleaning & validating…")
        try:
            df = clean_raw_file(filepath, log_fn=log_fn)
        except Exception as e:
            # Last resort: read as-is and patch missing columns
            if log_fn: log_fn(f"  Fallback to raw read ({e})")
            df = pd.read_excel(filepath, dtype=str)
            df = _normalise_columns(df)
            df["Date"] = df["Date"].apply(_clean_date)
            df["Time"] = df["Time"].apply(_clean_time)
            df["Class"] = df["Class"].apply(_clean_class) if "Class" in df.columns else pd.NA
            df["Type"]  = df["Type"].apply(_clean_type)   if "Type"  in df.columns else pd.NA
            df = df.dropna(subset=["Date", "Time"])
            df = df.sort_values(["Date", "Time"]).reset_index(drop=True)
            for col in CLEAN_COLS:
                if col not in df.columns:
                    df[col] = pd.NA
            df = df[CLEAN_COLS]

    if log_fn: log_fn(f"Cleaned dataset: {len(df):,} rows")

    if len(start_t) == 5: start_t += ":00"
    if len(end_t)   == 5: end_t   += ":59"

    mask     = df["Time"].apply(lambda t: start_t <= str(t) <= end_t)
    filtered = df[mask].copy().sort_values("Time").reset_index(drop=True)
    alarms   = filtered[filtered["Type"] == "alarm"].copy()

    if log_fn: log_fn(f"Filtered to {start_t[:5]}–{end_t[:5]}: {len(filtered):,} rows, {len(alarms):,} alarms")
    return filtered, alarms, df


def build_output_paths(input_file, start_t, end_t, outdir):
    stem   = Path(input_file).stem
    ts_s   = start_t[:5].replace(":", "_")
    ts_e   = end_t[:5].replace(":", "_")
    out_dir = Path(outdir)
    out_dir.mkdir(parents=True, exist_ok=True)
    return (str(out_dir / f"{stem}__filtered_{ts_s}_to_{ts_e}.xlsx"),
            str(out_dir / f"{stem}__report_{ts_s}_to_{ts_e}.xlsx"))


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 3 — GUI
# ══════════════════════════════════════════════════════════════════════════════

BG   = "#0f1923"; BG2  = "#162130"; BG3  = "#1e2d3d"
BDR  = "#2a3f54"; ACC  = "#2e75b6"; ACC2 = "#4472c4"
OK   = "#1db954"; WARN = "#f39c12"; ERR  = "#e74c3c"
TXT  = "#e8edf2"; DIM  = "#7f9ab0"; WHT  = "#ffffff"
FF   = "Segoe UI" if sys.platform == "win32" else "Arial"

CLS_FG = {"A":"#ff6b6b","B":"#f39c12","C":"#f1c40f","D":"#5dade2"}


def _btn(parent, text, cmd, kind="primary"):
    colors = {"primary":(ACC,WHT),"success":(OK,WHT),"danger":(ERR,WHT),"ghost":(BG3,DIM)}
    bg, fg = colors.get(kind, (ACC, WHT))
    b = tk.Button(parent, text=text, command=cmd, bg=bg, fg=fg,
                  activebackground=bg, activeforeground=fg,
                  relief="flat", bd=0, cursor="hand2",
                  font=(FF,10,"bold"), padx=14, pady=7)
    return b


class Card(tk.Frame):
    def __init__(self, p, **kw):
        super().__init__(p, bg=BG2, highlightbackground=BDR,
                         highlightthickness=1, **kw)


class TimePicker(tk.Frame):
    def __init__(self, parent, label):
        super().__init__(parent, bg=BG2)
        tk.Label(self, text=label, bg=BG2, fg=DIM, font=(FF,9)).pack(anchor="w", pady=(0,4))
        row = tk.Frame(self, bg=BG2)
        row.pack()
        self.hv = tk.StringVar(value="00")
        self.mv = tk.StringVar(value="00")
        sp = dict(bg=BG3, fg=TXT, insertbackground=TXT, relief="flat", bd=0,
                  highlightthickness=1, highlightbackground=BDR, highlightcolor=ACC,
                  font=(FF,14,"bold"), width=3, justify="center", buttonbackground=BG3)
        tk.Spinbox(row, from_=0, to=23, textvariable=self.hv,
                   format="%02.0f", **sp).pack(side="left")
        tk.Label(row, text=":", bg=BG2, fg=ACC, font=(FF,16,"bold")).pack(side="left", padx=4)
        tk.Spinbox(row, from_=0, to=59, textvariable=self.mv,
                   format="%02.0f", **sp).pack(side="left")

    def get(self): return f"{self.hv.get().zfill(2)}:{self.mv.get().zfill(2)}"
    def set(self, v):
        p = v.split(":")
        if len(p)>=2: self.hv.set(p[0].zfill(2)); self.mv.set(p[1].zfill(2))


class StatBadge(tk.Frame):
    def __init__(self, parent, label, color=ACC):
        super().__init__(parent, bg=BG2)
        self._v = tk.StringVar(value="—")
        tk.Label(self, textvariable=self._v, bg=BG2, fg=color,
                 font=(FF,20,"bold")).pack()
        tk.Label(self, text=label, bg=BG2, fg=DIM, font=(FF,8)).pack()
    def set(self, v): self._v.set(str(v))


class GlowBadge(tk.Frame):
    """Stat card with a coloured accent bar on the left."""
    def __init__(self, parent, label, icon="", color=ACC, width=130):
        super().__init__(parent, bg="#1a2a3a", highlightbackground=color,
                         highlightthickness=1, width=width)
        self.pack_propagate(False)
        # left accent stripe
        tk.Frame(self, bg=color, width=4).pack(side="left", fill="y")
        inner = tk.Frame(self, bg="#1a2a3a")
        inner.pack(fill="both", expand=True, padx=8, pady=10)
        self._v = tk.StringVar(value="—")
        tk.Label(inner, textvariable=self._v, bg="#1a2a3a", fg=color,
                 font=(FF, 18, "bold")).pack(anchor="w")
        tk.Label(inner, text=f"{icon}  {label}" if icon else label,
                 bg="#1a2a3a", fg=DIM, font=(FF, 8)).pack(anchor="w")
    def set(self, v): self._v.set(str(v))


class SPUApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("SPU Log Analyzer")
        self.geometry("1120x760")
        self.minsize(960, 660)
        self.configure(bg=BG)

        self._fp      = tk.StringVar()
        self._outdir  = tk.StringVar(value=str(Path.home() / "SPU_Reports"))
        self._running = False

        # ttk styles (set once here before any widget)
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Blue.Horizontal.TProgressbar",
                         troughcolor="#0a1520", background=ACC,
                         lightcolor=ACC, darkcolor=ACC2)

        self._build_ui()
        self._center()

    # ─── Main layout ──────────────────────────────────────────────────────────

    def _build_ui(self):
        # ══ TOP HEADER BAR ════════════════════════════════════════════════════
        hdr = tk.Frame(self, bg="#0d1e2e", height=60)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)

        # Logo area
        logo_area = tk.Frame(hdr, bg="#0d1e2e")
        logo_area.pack(side="left", padx=20, pady=8)
        tk.Label(logo_area, text="⚡", bg="#0d1e2e", fg=ACC,
                 font=(FF, 20)).pack(side="left")
        title_area = tk.Frame(logo_area, bg="#0d1e2e")
        title_area.pack(side="left", padx=(8, 0))
        tk.Label(title_area, text="SPU Log Analyzer", bg="#0d1e2e", fg=WHT,
                 font=(FF, 14, "bold")).pack(anchor="w")
        tk.Label(title_area, text="Raw Input  →  Clean  →  Filter  →  Report",
                 bg="#0d1e2e", fg=DIM, font=(FF, 8)).pack(anchor="w")

        # Right: version pill
        pill = tk.Frame(hdr, bg=ACC2, padx=10, pady=3)
        pill.pack(side="right", padx=20, pady=18)
        tk.Label(pill, text="v2.0", bg=ACC2, fg=WHT, font=(FF, 8, "bold")).pack()

        # ══ BODY (left panel + right panel) ══════════════════════════════════
        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True)

        # ── LEFT: Controls (fixed width sidebar) ──────────────────────────────
        sidebar = tk.Frame(body, bg="#0d1e2e", width=380)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)

        # thin accent line between sidebar and main
        tk.Frame(body, bg=ACC, width=2).pack(side="left", fill="y")

        # ── RIGHT: Stats + Log ────────────────────────────────────────────────
        main = tk.Frame(body, bg=BG)
        main.pack(side="left", fill="both", expand=True)

        self._build_sidebar(sidebar)
        self._build_main(main)

    # ─── Sidebar ──────────────────────────────────────────────────────────────

    def _build_sidebar(self, parent):
        # Scroll-capable canvas so sidebar content won't clip on small screens
        canvas = tk.Canvas(parent, bg="#0d1e2e", highlightthickness=0)
        sb     = tk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True)

        inner = tk.Frame(canvas, bg="#0d1e2e")
        win   = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _resize(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfig(win, width=e.width)
        inner.bind("<Configure>", _resize)
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(win, width=e.width))

        pad = {"padx": 18, "pady": (0, 14)}

        # ── Section: Input File ───────────────────────────────────────────────
        self._sb_section(inner, "INPUT FILE")
        fc = tk.Frame(inner, bg="#0d1e2e")
        fc.pack(fill="x", **pad)

        self._filetype_var = tk.StringVar(value="No file selected")
        self._file_entry = tk.Entry(fc, textvariable=self._fp,
                                     bg=BG3, fg=TXT, insertbackground=TXT,
                                     relief="flat", highlightthickness=1,
                                     highlightbackground=BDR, highlightcolor=ACC,
                                     font=(FF, 9))
        self._file_entry.pack(fill="x", ipady=8, pady=(0, 6))

        btn_row = tk.Frame(fc, bg="#0d1e2e")
        btn_row.pack(fill="x")
        browse_btn = _btn(btn_row, "📂  Browse File", self._browse, "primary")
        browse_btn.configure(pady=8, font=(FF, 9, "bold"))
        browse_btn.pack(fill="x")

        self._file_pill = tk.Label(fc, textvariable=self._filetype_var,
                                    bg="#0d1e2e", fg=DIM, font=(FF, 8),
                                    anchor="w")
        self._file_pill.pack(anchor="w", pady=(5, 0))

        # ── Section: Time Range ───────────────────────────────────────────────
        self._sb_section(inner, "TIME RANGE")
        tr = tk.Frame(inner, bg="#0d1e2e")
        tr.pack(fill="x", **pad)

        # Start time
        tk.Label(tr, text="Start Time", bg="#0d1e2e", fg=DIM,
                 font=(FF, 8)).pack(anchor="w", pady=(0, 3))
        self._start_p = TimePicker(tr, "")
        self._start_p.configure(bg="#0d1e2e")
        self._start_p.pack(anchor="w", pady=(0, 10))

        # Arrow
        tk.Label(tr, text="↓", bg="#0d1e2e", fg=ACC,
                 font=(FF, 14, "bold")).pack(anchor="w", pady=(0, 6))

        # End time
        tk.Label(tr, text="End Time", bg="#0d1e2e", fg=DIM,
                 font=(FF, 8)).pack(anchor="w", pady=(0, 3))
        self._end_p = TimePicker(tr, "")
        self._end_p.configure(bg="#0d1e2e")
        self._end_p.set("23:59")
        self._end_p.pack(anchor="w")

        # ── Section: Output ───────────────────────────────────────────────────
        self._sb_section(inner, "OUTPUT DIRECTORY")
        oc = tk.Frame(inner, bg="#0d1e2e")
        oc.pack(fill="x", **pad)

        tk.Entry(oc, textvariable=self._outdir,
                 bg=BG3, fg=TXT, insertbackground=TXT,
                 relief="flat", highlightthickness=1,
                 highlightbackground=BDR, highlightcolor=ACC,
                 font=(FF, 9)).pack(fill="x", ipady=8, pady=(0, 6))
        ob = _btn(oc, "📁  Choose Folder", self._browse_out, "ghost")
        ob.configure(pady=7, font=(FF, 9, "bold"))
        ob.pack(fill="x")

        # ── GENERATE BUTTON ───────────────────────────────────────────────────
        sep = tk.Frame(inner, bg=BDR, height=1)
        sep.pack(fill="x", padx=18, pady=(8, 16))

        btn_area = tk.Frame(inner, bg="#0d1e2e")
        btn_area.pack(fill="x", padx=18, pady=(0, 8))

        self._run_btn = tk.Button(btn_area, text="▶   Generate Reports",
                                   command=self._run,
                                   bg=OK, fg=WHT,
                                   activebackground="#17a844",
                                   activeforeground=WHT,
                                   relief="flat", bd=0, cursor="hand2",
                                   font=(FF, 12, "bold"), pady=13)
        self._run_btn.pack(fill="x", pady=(0, 8))

        self._open_btn = _btn(btn_area, "📂  Open Output Folder",
                               self._open_folder, "ghost")
        self._open_btn.configure(state="disabled", pady=8, font=(FF, 9))
        self._open_btn.pack(fill="x")

        # ── Progress + status ─────────────────────────────────────────────────
        pf = tk.Frame(inner, bg="#0d1e2e")
        pf.pack(fill="x", padx=18, pady=(10, 0))

        self._status = tk.StringVar(value="Ready — select a file to begin.")
        tk.Label(pf, textvariable=self._status, bg="#0d1e2e", fg=DIM,
                 font=(FF, 8), anchor="w", wraplength=340,
                 justify="left").pack(fill="x", pady=(0, 5))
        self._progress = ttk.Progressbar(pf, mode="indeterminate",
                                          style="Blue.Horizontal.TProgressbar",
                                          length=340)
        self._progress.pack(fill="x")

        # ── Output file list ──────────────────────────────────────────────────
        self._sb_section(inner, "OUTPUT FILES")
        self._files_frame = tk.Frame(inner, bg="#0d1e2e")
        self._files_frame.pack(fill="x", padx=18, pady=(0, 18))
        tk.Label(self._files_frame, text="No files generated yet.",
                 bg="#0d1e2e", fg=DIM, font=(FF, 8)).pack(anchor="w")

    def _sb_section(self, parent, title):
        """Sidebar section label with accent underline."""
        frame = tk.Frame(parent, bg="#0d1e2e")
        frame.pack(fill="x", padx=18, pady=(18, 8))
        tk.Label(frame, text=title, bg="#0d1e2e", fg=ACC,
                 font=(FF, 8, "bold")).pack(anchor="w")
        tk.Frame(frame, bg=ACC2, height=1).pack(fill="x", pady=(3, 0))

    # ─── Main panel (stats + log) ─────────────────────────────────────────────

    def _build_main(self, parent):
        # ── STAT BADGES ROW ───────────────────────────────────────────────────
        stats_row = tk.Frame(parent, bg=BG)
        stats_row.pack(fill="x", padx=20, pady=(16, 12))

        badge_defs = [
            ("Total Entries",  "📋", ACC2,           "_s_total"),
            ("Alarms",         "🔔", WARN,            "_s_alarms"),
            ("Class A",        "🔴", CLS_FG["A"],     "_s_a"),
            ("Class B",        "🟠", CLS_FG["B"],     "_s_b"),
            ("Class C",        "🟡", CLS_FG["C"],     "_s_c"),
            ("Class D",        "🔵", CLS_FG["D"],     "_s_d"),
        ]
        for label, icon, color, attr in badge_defs:
            badge = GlowBadge(stats_row, label, icon, color)
            badge.pack(side="left", fill="y", expand=True, padx=(0, 8), ipady=4)
            setattr(self, attr, badge)

        # ── DIVIDER ───────────────────────────────────────────────────────────
        tk.Frame(parent, bg=BDR, height=1).pack(fill="x", padx=20)

        # ── LOG HEADER ────────────────────────────────────────────────────────
        log_hdr = tk.Frame(parent, bg=BG)
        log_hdr.pack(fill="x", padx=20, pady=(10, 6))
        tk.Label(log_hdr, text="🖥  Activity Log", bg=BG, fg=TXT,
                 font=(FF, 10, "bold")).pack(side="left")
        clr = _btn(log_hdr, "Clear", self._clear_log, "ghost")
        clr.configure(font=(FF, 8), padx=8, pady=3)
        clr.pack(side="right")

        # ── LOG BODY ──────────────────────────────────────────────────────────
        log_frame = tk.Frame(parent, bg="#0a1520",
                             highlightbackground=BDR, highlightthickness=1)
        log_frame.pack(fill="both", expand=True, padx=20, pady=(0, 16))

        self._log = tk.Text(log_frame, bg="#0a1520", fg=TXT,
                            insertbackground=TXT, relief="flat", bd=0,
                            font=("Consolas", 9), wrap="word",
                            state="disabled", highlightthickness=0,
                            padx=12, pady=10)
        sb = tk.Scrollbar(log_frame, command=self._log.yview,
                          bg="#0a1520", troughcolor="#0a1520")
        self._log.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self._log.pack(fill="both", expand=True)

        self._log.tag_configure("ok",    foreground=OK,   font=("Consolas", 9, "bold"))
        self._log.tag_configure("err",   foreground=ERR,  font=("Consolas", 9, "bold"))
        self._log.tag_configure("warn",  foreground=WARN)
        self._log.tag_configure("dim",   foreground=DIM)
        self._log.tag_configure("bold",  foreground=TXT,  font=("Consolas", 9, "bold"))
        self._log.tag_configure("hdr",   foreground=ACC,
                                font=("Consolas", 9, "bold"), spacing1=6, spacing3=6)
        self._log.tag_configure("step",  foreground=ACC2, font=("Consolas", 9, "bold"))

        # Welcome message
        self._log_write("SPU Log Analyzer ready.", "hdr")
        self._log_write("Select a log file and time range, then click Generate Reports.", "dim")

    # ─── Helpers ──────────────────────────────────────────────────────────────

    def _center(self):
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        x = (self.winfo_screenwidth()  - w) // 2
        y = (self.winfo_screenheight() - h) // 2
        self.geometry(f"+{x}+{y}")

    def _log_write(self, msg, tag=""):
        self._log.configure(state="normal")
        ts = datetime.now().strftime("%H:%M:%S")
        self._log.insert("end", f"[{ts}]  ", "dim")
        self._log.insert("end", msg + "\n", tag)
        self._log.see("end")
        self._log.configure(state="disabled")

    def _ui_log(self, msg, tag=""):
        self.after(0, lambda: self._log_write(msg, tag))

    def _clear_log(self):
        self._log.configure(state="normal")
        self._log.delete("1.0", "end")
        self._log.configure(state="disabled")

    def _update_file_list(self, paths):
        for w in self._files_frame.winfo_children():
            w.destroy()
        for path in paths:
            row = tk.Frame(self._files_frame, bg="#0d1e2e")
            row.pack(fill="x", pady=3)
            icon = "📊" if "report" in Path(path).name.lower() else "📋"
            tk.Label(row, text=f"{icon}  {Path(path).name}",
                     bg="#0d1e2e", fg=TXT, font=(FF, 8),
                     anchor="w", wraplength=220).pack(side="left", fill="x", expand=True)
            tk.Button(row, text="Open ↗",
                      command=lambda p=path: self._open_file(p),
                      bg=ACC, fg=WHT, relief="flat",
                      font=(FF, 7, "bold"), padx=6, pady=2,
                      cursor="hand2").pack(side="right")

    def _open_file(self, path):
        try:
            if sys.platform == "win32":    os.startfile(path)
            elif sys.platform == "darwin": subprocess.Popen(["open", path])
            else:                          subprocess.Popen(["xdg-open", path])
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _open_folder(self):
        d = self._outdir.get()
        if os.path.isdir(d): self._open_file(d)
        else: messagebox.showwarning("Not found", f"Folder not found:\n{d}")

    # ─── Pickers ──────────────────────────────────────────────────────────────

    def _browse(self):
        p = filedialog.askopenfilename(
            title="Select SPU Log File",
            filetypes=[("SPU Log files", "*.csv *.xlsx *.xls *.txt *.log *.tsv"),
                       ("All files", "*.*")])
        if p:
            self._fp.set(p)
            ext = Path(p).suffix.upper()
            is_raw = ext in (".CSV", ".TXT", ".LOG", ".TSV")
            self._filetype_var.set(
                "⚡ Raw CSV — will be auto-cleaned" if is_raw
                else "✔ Excel file — will be validated")
            self._file_pill.configure(fg=WARN if is_raw else OK)
            self._log_write(f"Selected: {Path(p).name}", "dim")

    def _browse_out(self):
        p = filedialog.askdirectory(title="Select Output Directory")
        if p: self._outdir.set(p)

    # ─── Validation ───────────────────────────────────────────────────────────

    def _validate_time(self, s):
        s = s.strip()
        if re.fullmatch(r"\d{1,2}:\d{2}(:\d{2})?", s):
            p = s.split(":")
            if 0 <= int(p[0]) <= 23 and 0 <= int(p[1]) <= 59:
                return f"{int(p[0]):02d}:{int(p[1]):02d}"
        raise ValueError(f"Invalid time '{s}'")

    # ─── Run ──────────────────────────────────────────────────────────────────

    def _run(self):
        if self._running: return

        fp = self._fp.get().strip()
        if not fp or not os.path.isfile(fp):
            messagebox.showerror("Missing File", "Please select a valid log file.")
            return
        try:
            start = self._validate_time(self._start_p.get())
            end   = self._validate_time(self._end_p.get())
        except ValueError as e:
            messagebox.showerror("Invalid Time", str(e)); return
        if start >= end:
            messagebox.showerror("Invalid Range", "Start time must be before End time.")
            return

        outdir = self._outdir.get().strip() or "./output"
        self._running = True
        self._run_btn.configure(state="disabled", text="⏳  Working…",
                                bg="#555", activebackground="#555")
        self._open_btn.configure(state="disabled")
        self._progress.start(10)
        self._status.set("Processing — please wait…")

        self._log_write("━" * 48, "hdr")
        self._log_write(f"File:   {Path(fp).name}", "bold")
        self._log_write(f"Range:  {start} → {end}", "dim")
        self._log_write(f"Output: {outdir}", "dim")
        self._log_write("━" * 48, "hdr")

        threading.Thread(target=self._worker,
                         args=(fp, start, end, outdir), daemon=True).start()

    def _worker(self, fp, start, end, outdir):
        try:
            start_full = start + ":00"
            end_full   = end   + ":59"

            self._ui_log("Step 1 / 4  —  Loading & cleaning data…", "step")
            filtered, alarms, _ = load_and_filter(
                fp, start_full, end_full,
                log_fn=lambda m: self._ui_log("    " + m, "dim"))

            if len(filtered) == 0:
                self._ui_log(f"⚠  No entries found between {start} and {end}.", "warn")
                self._ui_log("   The file may not contain data in this window.", "warn")
                self._ui_done(False); return

            alarm_count = len(alarms)
            ca = len(alarms[alarms["Class"] == "A"]) if "Class" in alarms.columns else 0
            cb = len(alarms[alarms["Class"] == "B"]) if "Class" in alarms.columns else 0
            cc = len(alarms[alarms["Class"] == "C"]) if "Class" in alarms.columns else 0
            cd = len(alarms[alarms["Class"] == "D"]) if "Class" in alarms.columns else 0

            self._ui_log(f"✔  {len(filtered):,} entries  |  {alarm_count:,} alarms", "ok")
            if ca: self._ui_log(f"    🔴 Class A (Critical) : {ca}", "err")
            if cb: self._ui_log(f"    🟠 Class B (Major)    : {cb}", "warn")
            if cc: self._ui_log(f"    🟡 Class C (Minor)    : {cc}", "dim")
            if cd: self._ui_log(f"    🔵 Class D (Info)     : {cd}", "dim")

            self.after(0, lambda: [
                self._s_total.set(f"{len(filtered):,}"),
                self._s_alarms.set(f"{alarm_count:,}"),
                self._s_a.set(ca), self._s_b.set(cb),
                self._s_c.set(cc), self._s_d.set(cd)])

            out_data, out_report = build_output_paths(fp, start, end, outdir)

            self._ui_log("Step 2 / 4  —  Checking for locked files…", "step")
            for p in [out_data, out_report]:
                if os.path.exists(p):
                    try:
                        with open(p, "a"): pass
                    except PermissionError:
                        self._ui_log(f"⚠  File locked by Excel — please close:", "warn")
                        self._ui_log(f"    {Path(p).name}", "warn")
                        self._ui_done(False); return

            self._ui_log("Step 3 / 4  —  Generating filtered export…", "step")
            _sheet_filtered_export(filtered, alarms, start_full, end_full, out_data)
            self._ui_log("✔  Filtered export saved", "ok")

            self._ui_log("Step 4 / 4  —  Generating analysis report…", "step")
            build_analysis_report(filtered, alarms, start_full, end_full, out_report)
            self._ui_log("✔  Analysis report saved", "ok")

            self._ui_log("━" * 48, "hdr")
            self._ui_log("🎉  Done! Both files are ready.", "ok")

            self.after(0, lambda: self._update_file_list([out_data, out_report]))
            self._ui_done(True)

        except Exception as e:
            self._ui_log(f"✖  Error: {e}", "err")
            self._ui_log(traceback.format_exc(), "dim")
            self._ui_done(False)

    def _ui_done(self, success):
        def _upd():
            self._running = False
            self._progress.stop()
            self._run_btn.configure(state="normal",
                                    text="▶   Generate Reports",
                                    bg=OK, activebackground="#17a844")
            if success:
                self._status.set("✔  Reports generated successfully.")
                self._open_btn.configure(state="normal")
            else:
                self._status.set("✖  Failed — see Activity Log for details.")
        self.after(0, _upd)


# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    app = SPUApp()
    app.mainloop()
